#!/usr/bin/python2.7 
# -*- coding: utf-8 -*-  

import os, re, pdb
import time, datetime
import MySQLdb
from sqlalchemy import *
from sqlalchemy.orm import sessionmaker
from sqlalchemy.dialects.mysql import DOUBLE, TIMESTAMP
import numpy as np
import pandas as pd
import xlrd, openpyxl
from openpyxl.utils import get_column_letter

import DataLoggingModel

## Mysql Type
    # BIGINT, BINARY, BIT, BLOB, BOOLEAN, CHAR, DATE, \
    # DATETIME, DECIMAL, DECIMAL, DOUBLE, ENUM, FLOAT, INTEGER, \
    # LONGBLOB, LONGTEXT, MEDIUMBLOB, MEDIUMINT, MEDIUMTEXT, NCHAR, \
    # NUMERIC, NVARCHAR, REAL, SET, SMALLINT, TEXT, TIME, TIMESTAMP, \
    # TINYBLOB, TINYINT, TINYTEXT, VARBINARY, VARCHAR, YEAR
##

class DataBuildModel:
    def __init__(self, path = '', logPath = ''):
        if path == '':
            pass
        elif not os.path.exists(path):
            os.makedirs(path)
        self.path = path
        self.logger = DataLoggingModel.initLogger(logPath)

    def mysqlConnect(self, sqlCfg, use_SQLAlchemy = True):
        if 'host' in sqlCfg:
            _host = sqlCfg['host']
        else:
            _host = '127.0.0.1'
        self._host = _host

        if 'user' in sqlCfg:
            _user = sqlCfg['user']
        else:
            _user = 'root'
        self._user = _user

        if 'pwd' in sqlCfg:
            _pwd = sqlCfg['pwd']
        else:
            _pwd = ''
        self._pwd = _pwd
       
        if 'db' in sqlCfg:
            _db = sqlCfg['db']
        self._db = _db

        if 'port' in sqlCfg:
            _port = sqlCfg['port']
        else:
            _port = 3306
        self._port = _port

        if use_SQLAlchemy == True: # 使用其他引擎连接，仅初始化配置
            try:
                self.engine = create_engine('mysql://' + self._user + ':' + self._pwd + '@' + self._host + '/' + self._db + '?charset=utf8')
            except Exception, e:
                print 'Mysql Error (create engine):', e
        try:
            self.conn = MySQLdb.connect(host=_host, user=_user, passwd=_pwd, port=_port, charset='utf8')
            self.cursor = self.conn.cursor()
            self.conn.select_db(_db)
            print 'Mysql connected %s %s' % (_host, _db)
        except MySQLdb.Error,e:
            print 'Mysql Error %d: %s' % (e.args[0], e.args[1])

    def mysqlQuery(self, query = ''):
##        self.cursor.execute(query)
##        data = self.cursor.fetchone()
##        print data
        df = pd.read_sql(query, self.conn)
        return df

    def mysqlDestroy(self):
        try:
            self.cursor.close()
            self.conn.close()
            print 'Mysql destroyed %s %s' % (self._host, self._db)
        except MySQLdb.Error,e:
            print 'Mysql Error %d: %s' % (e.args[0], e.args[1])

    # if_exists: append/append_ignore/replace/fail/delete
    #       need_datetime: 是否需要额外追加created_at、updated_at, unique_key有值时生效
    def mysqlWriter(self, df, tb_name, sql_db, if_exists = 'fail', unique_key = None, need_datetime = True, need_connect = True):
        # engine = create_engine('mysql://' + self._user + ':' + self._pwd + '@' + self._host + '/' + self._db + '?charset=utf8')
        if need_connect:
            self.mysqlConnect(sql_db)

        if unique_key is None or unique_key == '':
            if if_exists == 'append_ignore':
                if_exists = 'append'
            if if_exists == 'delete':
                print 'Mysql ', tb_name, ' delete need compare keys'
                return
            df.to_sql(tb_name, self.engine, if_exists = if_exists)
            print 'Mysql ', tb_name, ' write succeed (No unique_key)!'

        # 包含有unique key的情况
        elif isinstance(unique_key, str) or isinstance(unique_key, unicode) or (isinstance(unique_key, list) and len(unique_key) >= 1):

            if isinstance(unique_key, str) or isinstance(unique_key, unicode):
                unique_key_list = [unique_key]
            elif isinstance(unique_key, list):
                unique_key_list = unique_key
            unique_key_map = dict()
            for ukey in unique_key_list:
                unique_key_map[ukey] = True

            # print self.cursor.fetchone()
            tb_count = self.cursor.execute('SHOW TABLES LIKE \"%' + tb_name + '%\";')
            tb_exists = False
            if tb_count > 0:
                tb_list = self.cursor.fetchall() # 获取数据表列表
                for tb_tmpname in tb_list:
                    if tb_name == tb_tmpname[0]: # 表已经存在
                        if if_exists == 'fail':
                            print 'table already exists and type is fail ... return'
                            return
                        elif if_exists == 'replace':
                            # self.cursor.execute('DROP TABLE IF EXISTS ' + tb_name + ';')
                            self.cursor.execute('DELETE FROM ' + tb_name + ';')
                            self.conn.commit()
                            print 'time to sleep 5 seconds'
                            time.sleep(5)
                            print 'clean table ', tb_name
                            tb_exists = True
                        elif if_exists == 'append' or if_exists == 'append_ignore' or if_exists == 'delete':
                            tb_exists = True

            if not tb_exists:   # 表不存在      
                created_at = False  # 判断是否存在此两个字段
                updated_at = False

                metadata = MetaData()
                colList = [tb_name, metadata, Column('id', BIGINT, primary_key = True)]
                # colList = [tb_name, metadata]

                for colName in df.columns:
                    colType = str(df[colName].dtype)
                    tmpType = String(128)
                    if colType.find('int') >= 0:
                        tmpType = Integer
                    elif colType.find('float') >= 0:
                        tmpType = DOUBLE
                    elif colType.find('datetime') >= 0:
                        tmpType = DateTime
                    else:
                    	if re.search(r'id$', colName):
                    	    tmpType = BIGINT
                    	elif re.search(r'(updated|created)', colName):
                    	    tmpType = TIMESTAMP
                    	else:
                            tmpType = String(128)
                    if colName in unique_key_map and len(unique_key_list) == 1:
                        colList.append(Column(colName, tmpType, unique = True, nullable = False, autoincrement = False)) 
                    elif colName in unique_key_map:
                        colList.append(Column(colName, tmpType, nullable = False)) 
                    else:
                        colList.append(Column(colName, tmpType))

                    if colName == 'created_at':
                        created_at = True
                    if colName == 'updated_at':
                        updated_at = True

                if need_datetime == True:
                    if updated_at == False:
                        colList.append(Column('updated_at', TIMESTAMP, nullable = False, server_default = text('CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP'))) 
                    if created_at == False:
                        colList.append(Column('created_at', TIMESTAMP, nullable = False))
                if len(unique_key_list) > 1:
                    colList.append(apply(UniqueConstraint, unique_key_list))
                tb_create = apply(Table, colList)   
                metadata.create_all(self.engine)
                print 'create table ', tb_name, ' succeed!'

            # 插入数据
            # DB_Session = sessionmaker(bind = self.engine)
            # session = DB_Session()  # sql orm

            metadata = MetaData(bind = self.engine)
            tb_model = Table(tb_name, metadata, autoload=True)

            tb_cols = dict()
            df_tmp = df.copy(deep = True)

            # 此处去除数据库中表不存在的列
            for colName in tb_model.c: 
                tb_cols[unicode(colName).replace(tb_name + '.', '')] = True
            for colName in df_tmp.columns:
                if unicode(colName) in tb_cols:
                    pass
                else:
                    df_tmp = df_tmp.drop(colName, axis=1)

            tb_count = df_tmp[unique_key_list[0]].count() # 数据记录条数
            for i in range(0, tb_count):
                dt_slice = df_tmp.iloc[i]
                insert_value = dt_slice.to_dict()

                if if_exists == 'replace':
                    dtime = datetime.datetime.now()
                    if 'updated_at' not in insert_value and 'updated_at' not in tb_cols and need_datetime:
                        insert_value['updated_at'] = dtime
                    if 'created_at' not in insert_value and 'created_at' not in tb_cols and need_datetime:
                        insert_value['created_at'] = dtime
                    try:
                        tb_model.insert(values = insert_value).execute()
                    except Exception, what: # 一般为 duplicate entry
                        if str(what).find('Duplicate entry') >= 0:
                            pass
                        else:
                            self.logger.error(str(insert_value) + ' |_| insert error ' + str(what))
                    continue

                else:
                    #首先查看数据是否存在
                    sql_query = 'SELECT COUNT(1) FROM ' + tb_name + ' WHERE '
                    tmp_list = []
                    for ukey in unique_key_list:
                        if need_datetime == True and (ukey == 'updated_at' or ukey == 'created_at'):
                            if ukey not in insert_value:
                                continue
                        if isinstance(dt_slice[ukey], unicode) or isinstance(dt_slice[ukey], str):
                            compare_str = dt_slice[ukey].replace('\"','\\\"')
                        else:
                            compare_str = str(dt_slice[ukey])
                        tmp_list.append('`' + ukey + '` = \"' + compare_str + '\"')
                    where_cause = ' AND '.join(tmp_list)    # 后面还会用
                    sql_query = sql_query + where_cause + ';'

                    dt_count = self.cursor.execute(sql_query)
                    dt_count = self.cursor.fetchone()
                    dt_count = dt_count[0]

                    # 涉及查询，位置不能换
                    dtime = datetime.datetime.now()
                    if 'updated_at' not in insert_value and 'updated_at' not in tb_cols and need_datetime:
                        insert_value['updated_at'] = dtime

                    if dt_count > 0:    # 记录已存在
                        if if_exists == 'delete':
                            try:
                                tb_model.delete().where(where_cause).execute()
                            except Exception, what:
                                # print 'mysql table ', tb_name, ' delete error: ', what
                                self.logger.error(str(insert_value) + ' |_| delete error ' + str(what))
                        elif if_exists == 'append_ignore':
                            continue
                        elif if_exists == 'append': # or if_exists == 'replace':   # 执行update
                            try:
                                # tb_model.update(values = insert_value).where(tb_model.c[ukey] == insert_value[ukey]).execute()
                                tb_model.update(values = insert_value).where(where_cause).execute()
                            except Exception, what:
                                self.logger.error(str(insert_value) + ' |_| update error ' + str(what))

                    elif if_exists != 'delete':   # 记录不存在，且type不为delete，执行insert
                        if 'created_at' not in insert_value and 'created_at' in tb_cols:
                            insert_value['created_at'] = dtime
                        try:
                            tb_model.insert(values = insert_value).execute()
                        except Exception, what: # 一般为 duplicate entry
                            if str(what).find('Duplicate entry') >= 0 and if_exists != 'append_ignore':
                                try:
                                    tb_model.update(values = insert_value).where(where_cause).execute()
                                except Exception, what:
                                    self.logger.error(str(insert_value) + ' |_| update error ' + str(what))
                            else:
                                self.logger.error(str(insert_value) + ' |_| insert error ' + str(what))

            # tb_model.execute() # 一起提交
            # session.close()
        else:
            print 'unique_key is not str or list'
        if need_connect:
            self.mysqlDestroy()     # 断开sql连接

    def csvReader(self, node = '1', fmt = True):
        if fmt:
            df = pd.read_csv(self.path + node + '.csv', index_col = 0, encoding='utf-8')
        else:
            df = pd.read_csv(self.path + node + '.csv', encoding='utf-8') #额外添加index_col
        return df

    def csvWriter(self, df, node = '1', encoding = 'utf-8'):
        df.to_csv(self.path + node + '.csv', encoding = encoding)
        # self.csvReEncoding(node)

    def csvReEncoding(self, node):
        df = self.csvReader(node, fmt = False)
        self.csvWriter(df, self.path + node + '_bak', 'gb2312')

    def excelReader(self, node = '1', sheet_name = 'Sheet1', fmt = True):
        if fmt:
            df = pd.read_excel(self.path + node + '.xlsx' , sheet_name, index_col = 0, header = 0 )
        else:
            df = pd.read_excel(self.path + node + '.xlsx' , sheet_name, index_col = None, header = 0 )
        return df            
    
    def excelWriter(self, df, node = '1', sheet_name = 'Sheet1', encoding = 'utf-8'):
        df.to_excel('%s%s.xlsx' % (self.path, unicode(node)), sheet_name = sheet_name, encoding = encoding)

    def excelWriterXls(self, df, node = '1', sheet_name = 'Sheet1', encoding = 'utf-8'):
        fileHandle = xlwt.Workbook()
        sheet1 = fileHandle.add_sheet(sheet_name, cell_overwrite_ok = True)
        
        dfIndexs = df.index
        dfColumns = df.columns
        dfValues = df.values
        for i in range(0, len(dfColumns)):
            sheet1.write(0, i + 1, dfColumns[i])
        for i in range(0, len(dfIndexs)):
            sheet1.write(i + 1, 0, dfIndexs[i])
            for j in range(0, len(dfColumns)):
                tmpValue = dfValues[i][j]
                # if isinstance(tmpValue, str):
                #     tmpValue = tmpValue.encoding('utf-8', 'ignore')
                #     if len(tmpValue) > 1000:
                #         tmpValue = tmpValue[0:1000]
                sheet1.write(i + 1, j + 1, tmpValue)
        fileHandle.save('%s%s.xls' % (self.path, unicode(node)))

    def excelWriterRaw(self, df, node = '1', if_exists = 'replace', sheet_name = 'Sheet1', encoding = 'utf-8'):
        if if_exists == 'append' or if_exists == 'append_ignore':
            wb = openpyxl.load_workbook(self.path + node + '.xlsx')
            ws = wb.get_sheet_by_name(sheet_name)
        elif if_exists == 'replace':
            wb = openpyxl.workbook.Workbook()
            ws = wb.create_sheet(sheet_name, 0)
        else:
            wb = openpyxl.workbook.Workbook()
            ws = wb.create_sheet(sheet_name, 0)
        
        dfIndexs = df.index
        dfColumns = df.columns
        dfValues = df.values

        for i in range(0, len(dfColumns)):
            index_col = i + 1
            index_row = 1
            col = get_column_letter(index_col + 1) 
            ws.cell( '%s%s' %(col, index_row)).value = dfColumns[i]
        for i in range(0, len(dfIndexs)):
            index_row = i + 2
            ws.cell('A'+str(index_row)).value = i           
            for j in range(0, len(dfColumns)):
                index_col = j + 1
                col = get_column_letter(index_col + 1) 
                tmpValue = dfValues[i][j]  
                try:
                    if (ws.cell('%s%s' %(col, index_row)).value is not None) and ws.cell('%s%s' %(col, index_row)).value != '':
                        if if_exists == 'append_ignore':
                            pass
                        elif if_exists == 'append' or if_exists == 'replace':
                            ws.cell( '%s%s' %(col, index_row)).value = str(tmpValue)
                    else:
                        ws.cell( '%s%s' %(col, index_row)).value = tmpValue
                    # ws['%s%s' %(col, index_row)] = int(tmpValue)
                except Exception, what:
                    # print 'encoding error: ', col, index_row, what
                    self.logger.error('write excel raw error: ' + str(col) + ' ' + str(index_row) + ' |_| reading error ' + str(what))
        wb.save(filename = '%s%s.xlsx' % (self.path, unicode(node)))

    def excelReEncoding(self, node):
        df = self.excelReader(node, fmt = False)
        self.excelWriter(df, node + '_bak', encoding = 'gb2312')


